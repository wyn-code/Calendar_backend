"""Exportación del calendario mensual a Excel (.xlsx) calcando el diseño de la web."""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.repositories.appointment_repository import AppointmentRepository

MESES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]
DIAS_SEMANA = ["DOM", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB"]

# --- Paleta exportada, sampleada de la UI real ---
COLOR_HEADER_BG = "8B6576"
COLOR_HEADER_FG = "FFFFFF"
COLOR_WEEKDAY_BG = "F9EDEB"
COLOR_WEEKDAY_FG = "6B4A57"
COLOR_DIA_BG = "FFFCFC"
COLOR_DIA_FG = "4C3F48"
COLOR_DIA_FUERA_BG = "F7F2F1"
COLOR_DIA_FUERA_FG = "B9AEB2"
COLOR_HOY_BG = "8B6576"
COLOR_HOY_FG = "FFFFFF"
COLOR_PARTICULAR_BG = "FED1D3"
COLOR_PARTICULAR_FG = "B8536F"
COLOR_OBRA_SOCIAL_BG = "DED6EF"
COLOR_OBRA_SOCIAL_FG = "6978C0"
COLOR_NOMBRE_FG = "7A73C0"
COLOR_HORA_FG = "2B2B3D"
COLOR_BORDE = "EDD9DC"


class CalendarioExcelService:
    """Construye el calendario mensual en Excel replicando la grilla de la web."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.repository = AppointmentRepository(db)

    def build(self, year: int, month: int) -> BytesIO:
        """Genera el workbook y devuelve el buffer .xlsx listo para descargar."""
        turnos_por_dia = self._turnos_por_dia(year, month)
        hoy = date.today()
        es_mes_actual = hoy.year == year and hoy.month == month

        wb = Workbook()
        ws = wb.active
        ws.title = f"{MESES[month - 1]} {year}"

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 24

        thin = Side(style="thin", color=COLOR_BORDE)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        self._escribir_header(ws, year, month)
        self._escribir_fila_dias(ws, border)

        fila = 3
        for dias_semana, max_turnos in self._semanas(year, month, turnos_por_dia):
            self._escribir_fila_numero_dia(
                ws, fila, border, dias_semana, month, es_mes_actual, hoy
            )
            fila += 1
            for slot in range(max_turnos):
                fila = self._escribir_fila_hora_nombre(
                    ws, fila, border, dias_semana, month, turnos_por_dia, slot
                )
                fila = self._escribir_fila_badge(
                    ws, fila, border, dias_semana, month, turnos_por_dia, slot
                )

        # Fija header + fila de días de semana al scrollear.
        ws.freeze_panes = "A4"

        # Orientación apaisada y ajuste a una hoja por si se imprime.
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _turnos_por_dia(
        self, year: int, month: int
    ) -> dict[date, list[tuple[str, str, str]]]:
        """Agrupa los turnos del mes por día: (hora, paciente, cobertura), ordenados por hora."""
        turnos: dict[date, list[tuple[str, str, str]]] = {}
        for appointment in self.repository.get_by_month(year, month):
            es_obra_social = appointment.tipo_consulta.strip().lower() == "obra social"
            if es_obra_social:
                cobertura = (
                    appointment.obra_social.nombre
                    if appointment.obra_social is not None
                    else "Obra Social"
                )
            else:
                cobertura = "Particular"
            turnos.setdefault(appointment.fecha, []).append(
                (
                    appointment.hora_inicio.strftime("%H:%M"),
                    appointment.patient.nombre_completo,
                    cobertura,
                )
            )
        for lista in turnos.values():
            lista.sort(key=lambda t: t[0])
        return turnos

    def _semanas(
        self, year: int, month: int, turnos_por_dia: dict[date, list[tuple[str, str, str]]]
    ) -> list[tuple[list[date], int]]:
        """6 semanas (DOM→SAB) que cubren el mes, con el máximo de turnos por semana."""
        primer_dia = date(year, month, 1)
        inicio = primer_dia - timedelta(days=(primer_dia.weekday() + 1) % 7)
        semanas: list[tuple[list[date], int]] = []
        for semana in range(6):
            dias = [inicio + timedelta(days=semana * 7 + d) for d in range(7)]
            # Mínimo 2 filas de turno (4 filas de Excel) por semana para que las
            # semanas sin turnos no se vean aplastadas al lado de semanas cargadas.
            max_turnos = max(2, *(len(turnos_por_dia.get(d, [])) for d in dias))
            semanas.append((dias, max_turnos))
        return semanas

    def _escribir_header(self, ws: Worksheet, year: int, month: int) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        cell = ws.cell(row=1, column=1, value=f"{MESES[month - 1].upper()} {year}")
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=20)
        cell.fill = self._fondo(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

    def _escribir_fila_dias(self, ws: Worksheet, border: Border) -> None:
        for col, dia in enumerate(DIAS_SEMANA, start=1):
            cell = ws.cell(row=2, column=col, value=dia)
            cell.font = Font(bold=True, color=COLOR_WEEKDAY_FG)
            cell.fill = self._fondo(COLOR_WEEKDAY_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def _escribir_fila_numero_dia(
        self,
        ws: Worksheet,
        fila: int,
        border: Border,
        dias: list[date],
        month: int,
        es_mes_actual: bool,
        hoy: date,
    ) -> None:
        for col, dia in enumerate(dias, start=1):
            cell = ws.cell(row=fila, column=col, value=dia.day)
            cell.fill, cell.font = self._estilo_dia(dia, month, es_mes_actual, hoy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[fila].height = 20

    def _escribir_fila_hora_nombre(
        self,
        ws: Worksheet,
        fila: int,
        border: Border,
        dias: list[date],
        month: int,
        turnos_por_dia: dict[date, list[tuple[str, str, str]]],
        slot: int,
    ) -> int:
        for col, dia in enumerate(dias, start=1):
            cell = ws.cell(row=fila, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.fill = self._fondo(COLOR_DIA_BG if dia.month == month else COLOR_DIA_FUERA_BG)
            lista = turnos_por_dia.get(dia, [])
            if slot < len(lista):
                hora, paciente, _ = lista[slot]
                cell.value = CellRichText(
                    TextBlock(InlineFont(b=True, color=COLOR_HORA_FG, sz="10"), f"{hora}  "),
                    TextBlock(InlineFont(b=False, color=COLOR_NOMBRE_FG, sz="10"), paciente),
                )
        return fila + 1

    def _escribir_fila_badge(
        self,
        ws: Worksheet,
        fila: int,
        border: Border,
        dias: list[date],
        month: int,
        turnos_por_dia: dict[date, list[tuple[str, str, str]]],
        slot: int,
    ) -> int:
        for col, dia in enumerate(dias, start=1):
            cell = ws.cell(row=fila, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.fill = self._fondo(COLOR_DIA_BG if dia.month == month else COLOR_DIA_FUERA_BG)
            lista = turnos_por_dia.get(dia, [])
            if slot < len(lista):
                _, _, cobertura = lista[slot]
                es_particular = cobertura == "Particular"
                cell.value = cobertura
                cell.fill = self._fondo(
                    COLOR_PARTICULAR_BG if es_particular else COLOR_OBRA_SOCIAL_BG
                )
                cell.font = Font(
                    bold=True,
                    color=COLOR_PARTICULAR_FG if es_particular else COLOR_OBRA_SOCIAL_FG,
                )
        return fila + 1

    def _fondo(self, color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=color)

    def _estilo_dia(
        self, dia: date, month: int, es_mes_actual: bool, hoy: date
    ) -> tuple[PatternFill, Font]:
        if es_mes_actual and dia == hoy:
            return self._fondo(COLOR_HOY_BG), Font(bold=True, color=COLOR_HOY_FG)
        if dia.month == month:
            return self._fondo(COLOR_DIA_BG), Font(bold=True, color=COLOR_DIA_FG)
        return self._fondo(COLOR_DIA_FUERA_BG), Font(bold=True, color=COLOR_DIA_FUERA_FG)
